#!/usr/bin/env python3
"""Add Hey Sister + Main Kitchen (Central Kitchen) station sheets to the
Station Training Template, matching the existing sheet format exactly.
Existing sheets are cloned (so styling/validation/structure is identical);
new content is only added on top. Nothing existing is removed."""
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

SRC = "/Users/mac/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/0D5F7783-740E-43FB-A26A-D6B19D6B7960/Station_Training_Template.xlsx"
OUT = "/Users/mac/Projects/MyMorAdmin/Station_Training_Template_UPDATED.xlsx"

wb = openpyxl.load_workbook(SRC)

CAT_DV = '"Opening,During Service,Closing,Weekly Clean,Monthly Deep Clean,Training - New Hire,Training - Ongoing,SOP Reference"'
SHIFT_DV = '"Morning,Night,All Day,N/A"'
PRIO_DV = '"High,Medium,Low"'
COMP_DV = '"Checkbox,Sign-off Required,Timer,Photo Required,N/A"'

# category value written into col C per section banner label
SECTION_CAT = {
    "OPENING": "Opening",
    "DURING SERVICE": "During Service",
    "CLOSING": "Closing",
    "WEEKLY CLEANING": "Weekly Clean",
    "MONTHLY DEEP CLEAN": "Monthly Deep Clean",
    "NEW HIRE TRAINING": "New Hire Training",
}


def readd_validations(ws):
    """Re-attach the 4 dropdowns over the data range (clone may drop them)."""
    ws.data_validations = DataValidationList()
    mr = ws.max_row
    for formula, col in [(CAT_DV, "C"), (SHIFT_DV, "E"), (PRIO_DV, "F"), (COMP_DV, "G")]:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.add(f"{col}5:{col}{mr}")
        ws.add_data_validation(dv)


def clone(src_title, new_title):
    new = wb.copy_worksheet(wb[src_title])
    new.title = new_title
    new.freeze_panes = "A5"
    readd_validations(new)
    return new


def set_banner(ws, text):
    ws["A1"] = text


def set_context(ws, shift_ctx, shared):
    ws["B2"] = shift_ctx   # B2:E2 merged
    ws["G2"] = shared      # G2:H2 merged


def section_rows(ws):
    """Map SECTION LABEL -> (banner_row, [data_row, ...]) by scanning col A."""
    banners = []
    for r in range(5, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper() in SECTION_CAT:
            banners.append((a.strip().upper(), r))
    out = {}
    for i, (label, br) in enumerate(banners):
        end = banners[i + 1][1] if i + 1 < len(banners) else ws.max_row + 1
        out[label] = (br, list(range(br + 1, end)))
    return out


def write_tasks(ws, content):
    """content = {SECTION_LABEL: [(task, notes, shift, priority, completion), ...]}"""
    secs = section_rows(ws)
    for label, items in content.items():
        if label not in secs:
            continue
        _, rows = secs[label]
        cat = SECTION_CAT[label]
        for item, r in zip(items, rows):
            task, notes, shift, prio, comp = item
            ws.cell(row=r, column=2, value=task)
            ws.cell(row=r, column=3, value=cat)
            ws.cell(row=r, column=4, value=notes)
            ws.cell(row=r, column=5, value=shift)
            ws.cell(row=r, column=6, value=prio)
            ws.cell(row=r, column=7, value=comp)


# ──────────────────────────────────────────────────────────────────────
# HEY SISTER — café like Mad Benji. Clone the MB equivalents (keeps their
# pre-populated tasks as reference) and re-label venue/station + context.
# (newtab, src, banner, shift_ctx, shared, area, index_desc)
# ──────────────────────────────────────────────────────────────────────
HS = [
    ("HS · FOH · Bar", "MB · FOH · Bar",
     "  Hey Sister  ·  FOH  ·  Bar (Counter & Help Floor)",
     "Morning — Bar helps Floor. Night — cold drinks and counter service.",
     "Floor (mutual cross-cover)", "FOH",
     "Morning — Bar helps Floor. Night — cold drinks & counter service."),
    ("HS · FOH · Barista", "MB · FOH · Barista",
     "  Hey Sister  ·  FOH  ·  Barista",
     "Barista helps Floor, Bar and Dishes across both shifts.",
     "Bar and Floor (cross-cover both shifts)", "FOH",
     "Barista helps Floor, Bar and Dishes across both shifts."),
    ("HS · FOH · Floor", "MB · FOH · Floor",
     "  Hey Sister  ·  FOH  ·  Floor",
     "Morning — Floor (help counter). Night — Prep & Floor.",
     "Bar (mutual cross-cover)", "FOH",
     "Morning — Floor (help counter). Night — Prep & Floor."),
    ("HS · BOH · Grill", "MB · BOH · Grill & Stove",
     "  Hey Sister  ·  BOH  ·  Grill",
     "Café kitchen grill — morning prep + service, dedicated at night.",
     "Fryer (shared equipment)", "BOH",
     "Café grill station — same structure as Mad Benji BOH."),
    ("HS · BOH · Fryer", "MB · BOH · Fryer",
     "  Hey Sister  ·  BOH  ·  Fryer",
     "Shared equipment (morning + night).",
     "Grill (shared equipment)", "BOH",
     "Shared equipment (morning + night)."),
    ("HS · BOH · Washing", "MB · BOH · Washing",
     "  Hey Sister  ·  BOH  ·  Washing",
     "Across both shifts.",
     "N/A", "BOH",
     "Across both shifts."),
    ("HS · BOH · Dressing", "MB · BOH · Dressing",
     "  Hey Sister  ·  BOH  ·  Dressing Bench",
     "Dressing / assembly bench (dedicated).",
     "N/A", "BOH",
     "Dressing / assembly bench (dedicated)."),
]

for newtab, src, banner, ctx, shared, area, desc in HS:
    ws = clone(src, newtab)
    set_banner(ws, banner)
    set_context(ws, ctx, shared)

# ──────────────────────────────────────────────────────────────────────
# MAIN KITCHEN — Central Kitchen (production/prep, not service). Clone the
# blank NEW STATION TEMPLATE skeleton and pre-populate from the HTML
# CK weekly prep schedule + standard production-kitchen tasks.
# ──────────────────────────────────────────────────────────────────────
TPL = "➕ NEW STATION TEMPLATE"

MK = [
    {
        "tab": "MK · CK · Prep & Cooking",
        "banner": "  Main Kitchen  ·  Central Kitchen  ·  Prep & Cooking",
        "ctx": "Production kitchen — runs the weekly prep cycle (Tue–Fri) and supplies all venues.",
        "shared": "Packaging & Distribution",
        "area": "CK",
        "desc": "Central production — weekly prep cycle (batch cooking) for all venues.",
        "content": {
            "OPENING": [
                ("Turn on ovens, steamers and prep equipment", "", "Morning", "High", "Checkbox"),
                ("Check & log fridge / freezer temperatures", "Record on the temp log", "Morning", "High", "Sign-off Required"),
                ("Review weekly prep schedule and the day's delivery list", "Tue/Wed/Thu/Fri each have a set run", "Morning", "High", "Checkbox"),
                ("Sanitise all prep benches and cutting boards", "", "Morning", "High", "Checkbox"),
            ],
            "DURING SERVICE": [
                ("Tuesday: Defrost prawns for dumplings; bake 30kg pumpkins", "", "All Day", "High", "Checkbox"),
                ("Tuesday: Drain chickpeas and collect dumplings", "", "All Day", "Medium", "Checkbox"),
                ("Tuesday: Make pumpkin patty mix and pumpkin hummus ×2", "", "All Day", "High", "Sign-off Required"),
                ("Wednesday: Make pumpkin patty at MB, bring back to CK", "", "All Day", "High", "Checkbox"),
                ("Thursday: Bake pulled beef FIRST (160°C, 20% steam, flip every 1.5 hrs)", "Start before anything else — long cook", "All Day", "High", "Timer"),
                ("Thursday: Make pork & prawn dumplings 20kg; cook bacon jam", "", "All Day", "High", "Sign-off Required"),
                ("Friday: Drain beef and mix with bacon jam", "", "All Day", "High", "Checkbox"),
            ],
            "CLOSING": [
                ("Label and date all prepared batches", "Product, date, destination venue", "All Day", "High", "Sign-off Required"),
                ("Store products in correct fridge / freezer zones", "", "All Day", "High", "Checkbox"),
                ("Turn off and clean ovens, steamers and mixers", "", "All Day", "High", "Checkbox"),
                ("Record production quantities in the log", "", "All Day", "Medium", "Sign-off Required"),
            ],
            "WEEKLY CLEANING": [
                ("Deep clean ovens and steamer trays", "", "All Day", "Medium", "Checkbox"),
            ],
            "NEW HIRE TRAINING": [
                ("Walk through the weekly prep schedule (Tue–Fri cycle)", "", "N/A", "High", "Sign-off Required"),
                ("Demonstrate batch recipes (pumpkin patty, dumplings, pulled beef)", "", "N/A", "High", "Checkbox"),
                ("Show labelling, dating and storage standards", "", "N/A", "High", "Checkbox"),
            ],
        },
    },
    {
        "tab": "MK · CK · Packaging & Dist",
        "banner": "  Main Kitchen  ·  Central Kitchen  ·  Packaging & Distribution",
        "ctx": "Portions, packs and delivers prepared product to Mad Benji, Hey Sister and Mad Hot Pot.",
        "shared": "Prep & Cooking",
        "area": "CK",
        "desc": "Portioning, packing and venue delivery runs.",
        "content": {
            "OPENING": [
                ("Check delivery run sheet and venue order quantities (MB, HS, MHP)", "", "Morning", "High", "Checkbox"),
                ("Prepare containers, labels and cold bags / Esky", "", "Morning", "High", "Checkbox"),
                ("Check van is clean and fuelled", "", "Morning", "Medium", "Checkbox"),
            ],
            "DURING SERVICE": [
                ("Portion and pack prepared items per venue order", "", "All Day", "High", "Checkbox"),
                ("Label each container with product, date and destination", "", "All Day", "High", "Sign-off Required"),
                ("Wednesday: Pick up van — collect beef (MB), stocks (HS), containers", "", "All Day", "High", "Checkbox"),
                ("Friday: Deliver beef + bacon jam to Hampton Park by 8:30am", "Time-critical delivery window", "All Day", "High", "Timer"),
                ("Load van in delivery order (last drop loaded first)", "", "All Day", "Medium", "Checkbox"),
            ],
            "CLOSING": [
                ("Confirm all venue orders dispatched and signed for", "", "All Day", "High", "Sign-off Required"),
                ("Return and store unused containers", "", "All Day", "Medium", "Checkbox"),
                ("Clean and sanitise van interior and cold bags", "", "All Day", "High", "Checkbox"),
            ],
            "NEW HIRE TRAINING": [
                ("Explain the venue delivery routes and timing windows", "", "N/A", "High", "Checkbox"),
                ("Show labelling and cold-chain requirements", "", "N/A", "High", "Checkbox"),
                ("Demonstrate correct van loading order", "", "N/A", "Medium", "Checkbox"),
            ],
        },
    },
    {
        "tab": "MK · CK · Cleaning & Maint",
        "banner": "  Main Kitchen  ·  Central Kitchen  ·  Cleaning & Maintenance",
        "ctx": "Keeps the production kitchen sanitised and equipment serviced across the prep day.",
        "shared": "Prep & Cooking",
        "area": "CK",
        "desc": "Sanitation and equipment maintenance for the production kitchen.",
        "content": {
            "OPENING": [
                ("Check chemical stock and set up dilution stations", "", "Morning", "High", "Checkbox"),
                ("Inspect equipment for faults; log any issues", "", "Morning", "High", "Sign-off Required"),
            ],
            "DURING SERVICE": [
                ("Wash and sanitise production utensils and containers between batches", "", "All Day", "High", "Checkbox"),
                ("Keep floors dry and clear during production", "", "All Day", "High", "Checkbox"),
                ("Empty and sanitise bins", "", "All Day", "Medium", "Checkbox"),
            ],
            "CLOSING": [
                ("Wash, rinse and sanitise all benches and equipment", "", "All Day", "High", "Checkbox"),
                ("Run final dishwasher cycle and put everything away", "", "All Day", "High", "Checkbox"),
                ("Sweep and mop the entire kitchen floor", "", "All Day", "High", "Checkbox"),
                ("Take out rubbish and recycling", "", "All Day", "Medium", "Checkbox"),
            ],
            "WEEKLY CLEANING": [
                ("Deep clean cool rooms and freezers", "", "All Day", "Medium", "Checkbox"),
                ("Descale dishwasher and sanitise drains", "", "All Day", "Medium", "Checkbox"),
            ],
            "MONTHLY DEEP CLEAN": [
                ("Full equipment strip-down clean (ovens, steamers, mixers)", "", "All Day", "High", "Photo Required"),
                ("Pest-control inspection and exhaust filter clean", "", "All Day", "High", "Sign-off Required"),
            ],
            "NEW HIRE TRAINING": [
                ("Chemical safety and dilution (SDS walkthrough)", "", "N/A", "High", "Sign-off Required"),
                ("Cleaning schedule and sign-off process", "", "N/A", "High", "Checkbox"),
            ],
        },
    },
]

REAL_A1_FILL = wb["MB · BOH · Grill & Stove"]["A1"].fill  # red banner from a real sheet
for spec in MK:
    ws = clone(TPL, spec["tab"])
    set_banner(ws, spec["banner"])
    # The blank template's top banner is grey; real sheets are red — match real.
    ws["A1"].fill = copy(REAL_A1_FILL)
    set_context(ws, spec["ctx"], spec["shared"])
    write_tasks(ws, spec["content"])

# ──────────────────────────────────────────────────────────────────────
# Reorder: new station sheets go before the NEW STATION TEMPLATE.
# ──────────────────────────────────────────────────────────────────────
order = [
    "📋 INDEX", "ℹ️ HOW TO USE",
    "MB · BOH · Grill & Stove", "MB · BOH · Pass", "MB · BOH · Fryer", "MB · BOH · Dressing", "MB · BOH · Washing",
    "MB · FOH · Bar", "MB · FOH · Floor", "MB · FOH · Barista",
    "MHP · FOH · Floor", "MHP · FOH · Counter", "MHP · BOH · Prep", "MHP · BOH · Boiler & Dishes", "MHP · BOH · Pass",
    "HS · FOH · Bar", "HS · FOH · Barista", "HS · FOH · Floor",
    "HS · BOH · Grill", "HS · BOH · Fryer", "HS · BOH · Washing", "HS · BOH · Dressing",
    "MK · CK · Prep & Cooking", "MK · CK · Packaging & Dist", "MK · CK · Cleaning & Maint",
    "➕ NEW STATION TEMPLATE",
]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)

# ──────────────────────────────────────────────────────────────────────
# INDEX: append the new stations (copy r5's per-column styling).
# ──────────────────────────────────────────────────────────────────────
ix = wb["📋 INDEX"]
style_row = 5  # donor for border/alignment/number_format
# Per-venue tint + text colour, mirroring MB (red) / MHP (green) in the file.
VENUE_STYLE = {
    "Hey Sister":   ("FDEFE2", "E67E22"),  # light orange / orange text
    "Main Kitchen": ("EAF1FD", "2563EB"),  # light blue / blue text
}
GRAY = "6B7280"
new_index = [(t, "Hey Sister", a, d) for (t, _s, _b, _c, _sh, a, d) in HS]
new_index += [(s["tab"], "Main Kitchen", s["area"], s["desc"]) for s in MK]

start = 18  # was the "➕ Add new station tab below" hint
for i, (tab, venue, area, desc) in enumerate(new_index):
    r = start + i
    fill_hex, text_hex = VENUE_STYLE[venue]
    fill = PatternFill(patternType="solid", fgColor=fill_hex)
    for col, val in [(2, tab), (3, venue), (4, area), (5, desc)]:
        src = ix.cell(row=style_row, column=col)
        c = ix.cell(row=r, column=col, value=val)
        # match the existing per-column pattern: B/C/D = venue colour (C bold),
        # E (description) = grey; everything else copied from the donor row.
        col_color = GRAY if col == 5 else text_hex
        c.font = Font(name=src.font.name, sz=src.font.sz, bold=src.font.bold, color=col_color)
        c.fill = fill
        c.alignment = copy(src.alignment)
        c.border = copy(src.border)
        c.number_format = src.number_format
    ix.row_dimensions[r].height = ix.row_dimensions[style_row].height
# keep the "add new station" hint at the bottom
hint_r = start + len(new_index)
src = ix.cell(row=4, column=2)
h = ix.cell(row=hint_r, column=2, value="➕  Add new station tab below")
h.font = Font(name=src.font.name, sz=src.font.sz, bold=False, color=GRAY)

wb.save(OUT)
print("saved", OUT)
print("total sheets:", len(wb.sheetnames))
print("new sheets:", [s for s in wb.sheetnames if s.startswith(("HS ", "MK "))])
