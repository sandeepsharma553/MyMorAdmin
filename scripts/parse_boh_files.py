#!/usr/bin/env python3
"""Parse the three BOH source files into per-file training (sectioned) + checklist
(flat) JSON for the restaurant-group platform. Emits to scripts/content/."""
import re, json, zipfile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

DL = "/Users/mac/Downloads/"
OUT = "/Users/mac/Projects/MyMorAdmin/scripts/content/"
W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

def clean(v):
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v).replace("☐", "").replace("☑", "")).strip()

def xlsx_lines(path):
    ws = load_workbook(DL + path, data_only=True)["Sheet1"]
    out = []
    for r in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in r if clean(c)]
        if cells:
            out.append(" ".join(cells))
    return out

def docx_lines(path):
    root = ET.fromstring(zipfile.ZipFile(DL + path).read("word/document.xml"))
    out = []
    for p in root.iter(W + "p"):
        t = clean("".join(t.text or "" for t in p.iter(W + "t")))
        if t:
            out.append(t)
    return out

def sectionize(lines, is_header, drop=()):
    """Split lines into [{heading, items}] using the is_header(line)->heading|None fn."""
    sections, cur = [], None
    for ln in lines:
        if ln in drop or not ln:
            continue
        h = is_header(ln)
        if h is not None:
            cur = {"heading": h, "items": []}
            sections.append(cur)
            continue
        if cur is None:
            cur = {"heading": "Procedure", "items": []}
            sections.append(cur)
        item = re.sub(r"^[-–•]\s*", "", ln).strip()
        if item:
            cur["items"].append(item)
    return [s for s in sections if s["items"]]

# numbered "1. Title" header for the two grill xlsx files
def numbered_header(title_line):
    def fn(ln):
        m = re.match(r"^\d+\.\s*(.+)$", ln)
        if m:
            return m.group(1).strip()
        return None
    return fn

SOUP_HEADERS = {"MAKE SOUP", "SIGNATURE", "LAKSA SOUP", "COLLAGEN SOUP", "TOMATO SOUP", "IMPERIAL SOUP"}
def hotpot_header(ln):
    u = re.sub(r"[,\(].*$", "", ln).strip().upper()
    if u in SOUP_HEADERS:
        return ln.strip().rstrip(",")
    return None

def build(src, title, key, venues, ctype, csub, sections):
    flat = [it for s in sections for it in s["items"]]
    desc = sections[0]["items"][0][:120] if sections and sections[0]["items"] else ""
    training = {"sourceTitle": src, "venues": venues, "modules": [{
        "key": key, "title": title, "cat": "BOH", "duration": "30 min", "icon": "🍳",
        "color": "#fde68a", "mandatory": True, "desc": desc, "steps": sections}]}
    checklist = {"sourceTitle": src, "venues": venues, "title": title, "type": ctype, "sub": csub, "items": flat, "key": key}
    json.dump(training, open(OUT + key + "-training.json", "w"), ensure_ascii=False, indent=2)
    json.dump(checklist, open(OUT + key + "-checklist.json", "w"), ensure_ascii=False, indent=2)
    print(f"{title}: {len(sections)} sections, {len(flat)} items → venues {venues}")

BOTH = ["hey-sister", "mad-hotpot"]

# 1) morning open boh — grill BOH opening
build("morning open boh.xlsx", "BOH Morning Open", "boh-morning-open", BOTH, "Opening", "Daily",
      sectionize(xlsx_lines("morning open boh.xlsx"), numbered_header("MORNING OPEN BOH"), drop=["MORNING OPEN BOH"]))

# 2) morning DUty — grill BOH duties/prep
build("morning DUty.xlsx", "BOH Morning Duty", "boh-morning-duty", BOTH, "Prep", "Daily",
      sectionize(xlsx_lines("morning DUty.xlsx"), numbered_header("Morning Duty Checklist"), drop=["Morning Duty Checklist"]))

# 3) Opening shop 1 — Mad Hot Pot opening + soups
build("Opening shop 1.docx", "Hot Pot Opening & Soups", "boh-hotpot-opening", ["mad-hotpot"], "Opening", "Daily",
      sectionize(docx_lines("Opening shop 1.docx"), hotpot_header))
