#!/usr/bin/env python3
"""Generate MyMor_Test_List.xlsx — a QA checklist covering Phases 1–5."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEAD_FILL = PatternFill("solid", fgColor="C0392B")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="F4D7D4")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("ID", 7),
    ("Feature / Area", 22),
    ("Test case", 40),
    ("Steps", 52),
    ("Expected result", 50),
    ("Test as role", 16),
    ("Pass / Fail", 12),
    ("Notes", 26),
]

def style_sheet(ws, rows):
    # header
    for c, (name, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    r = 2
    for row in rows:
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(row=r, column=c, value=row[c - 1] if c - 1 < len(row) else "")
            cell.alignment = WRAP
            cell.border = BORDER
            if c in (6, 7):
                cell.alignment = CENTER
        r += 1
    # auto-ish row height
    for rr in range(2, r):
        ws.row_dimensions[rr].height = 42

# ── Phase 1 — Quick wins ──
p1 = [
    ["1.1", "See passwords", "Reveal staff login credentials", "User Management → toggle 'Show logins'", "Email, password and POS PIN become visible for each staff with a login", "Store Admin", "", ""],
    ["1.2", "Staff end date", "Record a leaving date", "Staff Directory → open profile → Edit → set End date → Save", "End date saved and shown on profile; status can be set to Left", "Manager", "", ""],
    ["1.3", "Same-name staff", "Add two people with identical names", "Staff Directory → add 'Chirag Agarwal' twice", "Second one auto-named 'Chirag Agarwal 2'", "Store Admin", "", ""],
    ["1.4", "POS PIN", "Auto-generate a 4-digit PIN", "Add staff → click Auto next to PIN", "Unique 4-digit PIN generated; duplicates rejected on save", "Store Admin", "", ""],
    ["1.5", "Optional admin login", "Create staff without a website login", "Add staff, leave 'Admin website access' unchecked", "Staff saved with no email/password; PIN still works", "Store Admin", "", ""],
    ["1.6", "Daily checklist reset", "Checklist resets next day + history", "Checklists → tick items → use Reset / next day", "Ticks clear for the new day; prior day archived in History modal", "Manager", "", ""],
    ["1.7", "Training redirect link", "Open external training", "Training → module with link → open assignment → Open external training", "Opens the external URL in a new tab", "Staff", "", ""],
    ["1.8", "Shift click shows note", "View a shift's note", "Shift Planner → click a shift cell", "Detail modal opens with day/time/role/venue/notes (not instant delete)", "Manager", "", ""],
]

# ── Phase 2 — Stations & Settings ──
p2 = [
    ["2.1", "Settings page access", "Only admins see Settings", "Sign in as manager vs store admin", "Manager: no Settings nav; Store Admin/Owner: Settings visible", "Both", "", ""],
    ["2.2", "Create station", "Add a station to a venue", "Settings → Stations → pick venue → + Add station (e.g. Grill / BOH)", "Station appears in that venue's list with FOH/BOH pill", "Store Admin", "", ""],
    ["2.3", "Quick-add stations", "Use suggested stations", "Settings → Stations → Quick add buttons", "Clicking adds the station; button disables once added", "Store Admin", "", ""],
    ["2.4", "Per-venue stations", "Stations are venue-specific", "Add 'Grill' to Mad Benji only, switch venue tab", "Grill shows only under Mad Benji", "Store Admin", "", ""],
    ["2.5", "Roles CRUD", "Add/remove a role", "Settings → Roles → add 'Waitress' / remove one", "Role list updates; new role appears in staff & shift role pickers", "Store Admin", "", ""],
    ["2.6", "Assign stations to staff", "Pick stations on a profile", "Staff → add/edit → Stations picker (filtered by chosen venues)", "Only stations of the staff's venues show; saved on profile", "Store Admin", "", ""],
    ["2.7", "Station on shift", "Set a station for a shift", "Shift Planner → Add shift → choose Station", "Station saved; shows in shift cell & detail", "Manager", "", ""],
    ["2.8", "Station on checklist", "Tag a checklist with a station", "Checklists → edit → Station (optional)", "Station pill shows on the checklist card", "Manager", "", ""],
    ["2.9", "Station on training", "Tag a module with a station", "Training → edit module → Station (optional)", "Station pill shows on the module card", "Manager", "", ""],
]

# ── Phase 3 — Records & audit ──
p3 = [
    ["3.1", "Role/venue history", "Changes are logged on the profile", "Staff → edit role/venue/station/status → Confirm & save", "'Role & venue history' lists each change with date + who", "Store Admin", "", ""],
    ["3.2", "Change double-confirm", "Confirmation before saving", "Staff → edit → Save changes", "A confirm banner appears; must click 'Confirm & save'", "Store Admin", "", ""],
    ["3.3", "Coaching record", "Log a coaching note", "Staff profile → Coaching & mistake records → add (Coaching/Mistake/…)", "Entry added with type pill, note, date and author; removable", "Manager", "", ""],
    ["3.4", "Super-admin activity feed", "Sensitive changes notify super admin", "Super Admin → Restaurant Groups → Activity button", "Edits/permission changes/new staff appear; NEW badge + count", "Super Admin", "", ""],
    ["3.5", "Mark activity read", "Clear the NEW badge", "Activity modal → Mark all read", "Items lose NEW highlight; row badge count resets", "Super Admin", "", ""],
    ["3.6", "Permission change logged", "Permission edits are audited", "User Management → edit permissions → save", "Audit entry 'Permissions changed for …' appears in Activity", "Store Admin", "", ""],
]

# ── Phase 4 — Messaging ──
p4 = [
    ["4.1", "Post announcement", "Broadcast to all venues", "Messages → Announcements → write → All venues → Post", "Announcement appears for everyone in scope", "Manager", "", ""],
    ["4.2", "Venue-scoped announcement", "Target a single venue", "Messages → Announcements → choose a venue → Post", "Only staff of that venue see it", "Manager", "", ""],
    ["4.3", "Acknowledge", "Confirm read of an announcement", "Open an announcement → Acknowledge", "Shows ✓ Acknowledged; acknowledged count increments", "Staff", "", ""],
    ["4.4", "Direct message", "Two-way chat", "Messages → Direct → + New → pick a person → send", "Message appears in thread; recipient sees it in their list", "Manager", "", ""],
    ["4.5", "Unread badge", "Sidebar shows unread count", "Receive a DM / new announcement", "Messages nav shows a badge; clears when read", "Staff", "", ""],
    ["4.6", "Read receipts", "Opening a thread marks read", "Open a conversation with unread messages", "Unread count for that thread clears", "Manager", "", ""],
    ["4.7", "Permission gating", "Staff can read but not post announcements", "Sign in as staff", "Staff can view & acknowledge; compose box hidden", "Staff", "", ""],
]

# ── Phase 5 — Shifts, prep, trainer, notes ──
p5 = [
    ["5.1", "Trainer verifies training", "Trainer confirms competency", "Training/Staff → open assignment → Trainer verification → Verify", "Shows ✓ Verified by <name> + note; ✓ Verified pill on profile", "Manager (trainer)", "", ""],
    ["5.2", "Remove verification", "Undo a verification", "Open verified assignment → Remove verification", "Reverts to not verified", "Manager", "", ""],
    ["5.3", "Time-based checklist", "Set a scheduled time", "Checklists → edit → Scheduled time (e.g. 07:15)", "⏰ time pill shows; list sorted by time", "Manager", "", ""],
    ["5.4", "Due-now highlight", "Overdue checklist flagged", "Set a checklist time earlier than now, leave items unticked", "Card highlighted with 'Due now' pill", "Manager", "", ""],
    ["5.5", "Prep list carryover", "Items persist across days", "Checklists → Prep list → add item → next day", "Item still present (does NOT reset like checklists)", "Manager", "", ""],
    ["5.6", "Prep item note + qty", "Add detail to a prep item", "Prep list → add with Qty + Note / edit Note", "Qty (× n) and 📝 note shown under the item", "Staff", "", ""],
    ["5.7", "Reset prep ticks", "Clear done items for next service", "Prep list → tick some → Reset ticks", "All ticks clear; items remain", "Manager", "", ""],
    ["5.8", "Split-screen planner", "Compare two venues side by side", "Shift Planner → Split view → choose venue A & B", "Two rosters render side by side with per-venue hours", "Manager", "", ""],
    ["5.9", "Shift station in cell", "Station visible on roster", "Split/normal view → shift with a station", "Station shown under the time in the cell", "Manager", "", ""],
]

# ── Regression / cross-cutting ──
reg = [
    ["R.1", "Named database", "Data reads/writes hit mymor-australia", "Any create/edit; check Firebase console (mymor-australia)", "Docs land in the named DB, not (default)", "Store Admin", "", ""],
    ["R.2", "All-venues merge", "Selecting 'All venues' merges data", "Top venue filter → All venues", "Staff/checklists/training/shifts from every venue shown", "Owner", "", ""],
    ["R.3", "FOH/BOH filter", "Area filter incl. universal", "Checklists/Training → FOH / BOH tabs", "FOH shows FOH+All; BOH shows BOH+All", "Manager", "", ""],
    ["R.4", "Permission gating", "Nav hidden by permission", "Set a module to None for a role", "That nav item disappears for the role", "Store Admin", "", ""],
    ["R.5", "Rich text items", "Bold/colour/highlight persist", "Checklist/Training item → format with toolbar → save", "Formatting renders back correctly", "Manager", "", ""],
]

sheets = [
    ("Phase 1 - Quick wins", p1),
    ("Phase 2 - Stations", p2),
    ("Phase 3 - Records & audit", p3),
    ("Phase 4 - Messaging", p4),
    ("Phase 5 - Shifts & prep", p5),
    ("Regression", reg),
]

wb.remove(wb.active)
for name, rows in sheets:
    ws = wb.create_sheet(title=name[:31])
    style_sheet(ws, rows)

out = "MyMor_Test_List.xlsx"
wb.save(out)
print("wrote", out, "with", len(sheets), "sheets and", sum(len(r) for _, r in sheets), "test cases")
