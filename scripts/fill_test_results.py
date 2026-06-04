#!/usr/bin/env python3
"""Fill the Pass/Fail + Notes columns in MyMor_Test_List.xlsx with the
code-level audit verdicts (4 parallel agents, build verified)."""
import openpyxl
from openpyxl.styles import PatternFill, Font

GREEN = PatternFill("solid", fgColor="C6EFCE")
GREEN_F = Font(color="006100", bold=True)
AMBER = PatternFill("solid", fgColor="FFEB9C")
AMBER_F = Font(color="9C6500", bold=True)

# id -> (verdict, note)
V = {
    # Phase 1
    "1.1": ("PASS", "Show-logins toggle reveals email/password/PIN."),
    "1.2": ("PASS", "End date saved on add + edit, shown on profile."),
    "1.3": ("PASS", "uniqueDisplayName auto-suffixes ' 2'."),
    "1.4": ("PASS", "genPin excludes used PINs; duplicate rejected on save."),
    "1.5": ("PASS", "No login path works; PIN still generated."),
    "1.6": ("PASS", "checkedDate rollover + pushHist archive + History modal."),
    "1.7": ("PASS", "assignment.link opens new tab from AssignmentDetail."),
    "1.8": ("PASS", "Cell click opens detail modal; delete only inside modal."),
    # Phase 2
    "2.1": ("PASS", "FIXED during audit: page had no route guard (nav was hidden but URL rendered read-only). Added can('settings','view') guard."),
    "2.2": ("PASS", "saveStation writes {name,area,venueId,order}."),
    "2.3": ("PASS", "Quick-add disables once station exists."),
    "2.4": ("PASS", "Stations stored under venues/{id}/stations; all consumers filter by venueId."),
    "2.5": ("PASS", "Roles persisted on group doc; used in staff + shift pickers."),
    "2.6": ("PASS", "StationPicker filtered by venueIds; saves stationIds+stationNames."),
    "2.7": ("PASS", "Shift station select; saved + shown in cell & detail."),
    "2.8": ("PASS", "Checklist station select + card pill."),
    "2.9": ("PASS", "Training station select + card pill."),
    # Phase 3
    "3.1": ("PASS", "diffStaff() appends history via arrayUnion (ISO date, no serverTimestamp)."),
    "3.2": ("PASS", "Two-step Confirm & save + audit write."),
    "3.3": ("PASS", "records[] add/remove with unique id; newest-first."),
    "3.4": ("PASS", "Activity modal lists auditLog live; NEW badge + row count. Note: row count is one-time fetch, refreshes on reload."),
    "3.5": ("PASS", "Mark all read batch-sets seenBySuper + clears count."),
    "3.6": ("PASS", "Permission save writes perms.update audit entry."),
    # Phase 4
    "4.1": ("PASS", "postAnnouncement scope 'all'; gated by can('messages','edit')."),
    "4.2": ("PASS", "visibleAnns filters by myVenueIds for venue-scoped posts."),
    "4.3": ("PASS", "Acknowledge arrayUnion + count; re-ack guarded."),
    "4.4": ("PASS", "sendMessage writes conv=convId(); appears in thread."),
    "4.5": ("PASS", "unreadMessages badge = DMs + un-acked; own items pre-seeded read."),
    "4.6": ("PASS", "Opening thread marks incoming read; count clears."),
    "4.7": ("PASS", "Staff = view; compose hidden, acknowledge still shown. Identity match verified consistent."),
    # Phase 5
    "5.1": ("PASS", "Verify writes to correct trainingAssignments path; ✓ pill on profile."),
    "5.2": ("PASS", "Remove verification works (verifiedAt left stale — cosmetic)."),
    "5.3": ("PASS", "time field saved; list time-sorted."),
    "5.4": ("PASS", "Due-now: zero-padded HH:MM string compare correct."),
    "5.5": ("PASS", "Prep list has no daily-reset logic — carries over."),
    "5.6": ("PASS", "qty + editable note render with the item."),
    "5.7": ("PASS", "Reset ticks clears done, keeps items."),
    "5.8": ("PASS", "Split view = two VenueGrids, venue-scoped cells, per-venue hours."),
    "5.9": ("PASS", "FIXED during audit: main roster cell now shows station (was venue initials only)."),
    # Regression
    "R.1": ("PASS", "initializeFirestore named DB, not (default)."),
    "R.2": ("PASS", "flat() merges per-venue collections for 'All venues'."),
    "R.3": ("PASS", "FOH=FOH+All, BOH=BOH+All in checklists & training."),
    "R.4": ("PASS", "Nav filtered by can(key,'view'); 'none' hides it."),
    "R.5": ("PASS", "Rich text persisted as HTML and re-rendered."),
}

wb = openpyxl.load_workbook("MyMor_Test_List.xlsx")
filled = 0
for ws in wb.worksheets:
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(row=r, column=1).value
        if tid in V:
            verdict, note = V[tid]
            c7 = ws.cell(row=r, column=7, value=verdict)
            c8 = ws.cell(row=r, column=8, value=note)
            fixed = "FIXED" in note
            c7.fill = AMBER if fixed else GREEN
            c7.font = AMBER_F if fixed else GREEN_F
            filled += 1

wb.save("MyMor_Test_List.xlsx")
print(f"filled {filled} verdicts")
