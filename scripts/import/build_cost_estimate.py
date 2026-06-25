#!/usr/bin/env python3
"""
Build a DEFENSIBLE food/recipe cost ESTIMATE for every imported Mad Benji menu item.

READ-ONLY. Produces an .xlsx report only. Writes NOTHING to Firestore and populates
no live cost/recipe field. Every figure is an UNVERIFIED planning estimate derived
from researched AU 2025-2026 hospitality benchmarks, pending chef sign-off.

Inputs (local files):
  scripts/import/mymor_madbenji_import.json  (or the spaces-named variant) — menuItems
  scripts/import/_inventory_costs.json        — seed inventoryItems cost/unit (bottom-up)

Output:
  scripts/import/MadBenji_Menu_Cost_Estimate.xlsx
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

def load_menu():
    for name in ("mymor_madbenji_import.json",):
        p = os.path.join(HERE, name)
        if os.path.exists(p):
            with open(p, encoding="utf-8") as f:
                return json.load(f), p
    raise SystemExit("menu JSON not found in scripts/import/")

def load_inventory():
    p = os.path.join(HERE, "_inventory_costs.json")
    if not os.path.exists(p):
        return []
    with open(p, encoding="utf-8") as f:
        return json.load(f)

# ── Researched category benchmarks (AU 2025-2026 hospitality) ──────────────────
# classification: "Made in-house" | "Bought-in / resale" | "Unknown"
# low/mid/high are FOOD-COST % of EX-GST sell price. basis = sourced rationale.
MADE = "Made in-house"
RESALE = "Bought-in / resale"
UNK = "Unknown"
BENCHMARKS = {
    "Coffee & Tea":               (MADE,   18, 22, 28, "Espresso/tea made in-house. Coffee COGS ~20-28% of cup revenue incl. beans (~$0.18), milk (~$0.10), takeaway cup/lid (~$0.30); beverage GP 70-85% → 15-30% cost. [Fish River Roasters; Pool Six 2025]"),
    "Iced Drinks":                (MADE,   22, 27, 32, "Iced coffee/latte: milk + ice + cold cup raise cost vs hot; cafe cold-beverage cost ~25-30%."),
    "🥤 Shakes":                  (MADE,   25, 30, 38, "Milkshakes: ice cream + milk + syrup; blended-drink food cost ~25-38%."),
    "VEGAN Shakes":               (MADE,   28, 33, 40, "Plant milk + vegan ice cream cost ~3-5 pts more than dairy shakes."),
    "Smoothies":                  (MADE,   28, 35, 45, "Fresh-fruit smoothies + yoghurt/milk; fresh-ingredient smoothie food cost can reach ~50%. [Toast / juice-bar 2025]"),
    "VEGAN Smoothies":            (MADE,   28, 35, 45, "As smoothies; plant-based inputs comparable or higher."),
    "Juices":                     (MADE,   30, 38, 50, "Fresh-pressed juice: high produce input + yield loss; fresh juice ~40-60% cost. [Goodnature]"),
    "Sparkling Pops":             (MADE,   28, 35, 45, "⚠ ASSUMED house-made sparkling/soda. If bought-in packaged, reclassify as resale (35-50%)."),
    "Soft Drinks":                (RESALE, 35, 42, 50, "Packaged cans/bottles resold at low markup → high passthrough cost."),
    "Beer":                       (RESALE, 25, 32, 40, "Beer ~20-30% of sales (liquor benchmarks); AU cafe packaged at upper end."),
    "Beer Deal":                  (RESALE, 30, 38, 45, "Discounted beer bundle → higher cost% than list beer."),
    "🥧 Pastries":                (RESALE, 45, 50, 58, "Bought-in wholesale pastries: cafe wholesale 'barely breaks even' vs retail. [Baron Accounting AU]"),
    "Croissant":                  (RESALE, 45, 50, 58, "Bought-in baked goods; as pastries."),
    "Retails":                    (RESALE, 55, 65, 75, "⚠ Pure retail resale (packaged goods/merch). Thin markup, high passthrough. Wide price range ($5-$81) → mixed item types; recategorise."),
    "🍔 Burgers":                 (MADE,   35, 40, 45, "Gourmet burgers; premium brisket patty + bun + cheese; beef cost +32% since 2023. Bottom-up ~40-48%."),
    "Rolls":                      (MADE,   30, 35, 42, "Filled rolls; protein + bread base."),
    "Bagels":                     (MADE,   30, 35, 42, "Filled bagels; bought-in bagel + house fillings."),
    "Toasties":                   (MADE,   28, 33, 40, "Toasted sandwiches; bread + cheese/ham fillings."),
    "Breaky":                     (MADE,   30, 35, 42, "Cooked breakfasts (eggs/bacon/toast); cafe mid-high."),
    "Vegetarian & Vegan Breaky":  (MADE,   28, 34, 42, "Veg/vegan breakfasts; produce-heavy."),
    "Poke Bowls":                 (MADE,   32, 38, 45, "Raw fish/protein + rice + veg; protein-heavy."),
    "🥗Salads":                   (MADE,   28, 33, 40, "Composed salads; produce + protein."),
    "Loaded":                     (MADE,   28, 33, 40, "Loaded fries/nachos; cheap base, toppings raise cost."),
    "Bites":                      (MADE,   28, 33, 40, "Fried/snack bites; portioned protein/veg."),
    "🍟   Sides":                 (MADE,   22, 28, 35, "Chips/potato sides; low raw cost."),
    "Kids Menu":                  (MADE,   25, 30, 38, "Smaller portions; mixed."),
    "Dips & Sauce":               (MADE,   20, 28, 40, "⚠ Add-ons with tiny absolute price ($0.45-2.27); per-portion cost% highly volatile."),
    "Uncategorised":              (UNK,    25, 32, 40, "⚠ No category signal; uses overall AU cafe mid 32% (best-practice 25-28%, typical 28-35%). Recategorise."),
}
FALLBACK = (UNK, 25, 32, 40, "⚠ Category not in benchmark table; overall AU cafe mid 32% applied. Review.")
OVERALL_NOTE = "AU cafe/QSR food cost: best-practice 25-28%, typical 28-35% (Loaded Hub; VantaInsights). ATO coffee-shop blended cost-of-sales 34-40% of turnover."

SOURCES = [
    "Loaded Hub — Good food cost % for a restaurant in Australia: https://www.loadedhub.com/resources/good-food-cost-percentage-restaurant-australia",
    "VantaInsights — Restaurant food cost % 2026 (typical 28-35%): https://vantainsights.com/insights/restaurant-food-cost-percentage",
    "Fish River Roasters — Cost to make a cup of coffee (AU): https://www.fishriverroasters.com.au/blog/how-much-does-it-cost-to-make-a-cup-of-coffee",
    "Pool Six Coffee Roasters — Coffee shop margins 2025: https://blog.poolsixcoffeeroasters.com/setting-the-margins-on-your-coffee-menu/",
    "ATO — Small business benchmarks, coffee shops (cost of sales 34-40%): https://www.ato.gov.au/businesses-and-organisations/income-deductions-and-concessions/small-business-benchmarks/in-detail/coffee-shops",
    "Baron Accounting — Australian cake shops & patisseries economics (wholesale pastry margins): https://www.baronaccounting.com/post/a-financial-guide-for-australian-cake-shops-and-patisseries-industry-economics-performance",
    "Goodnature — Calculating food costs for juice / how much to charge: https://www.goodnature.com/blog/how-much-to-charge-for-juice",
    "Toast POS — Smoothie/juice bar cost guide & food cost %: https://pos.toasttab.com/blog/on-the-line/how-much-do-juice-bars-make",
    "Restaurant Dive — Beef costs +32% vs burger prices +14% (2023-2025): https://www.restaurantdive.com/news/Datassential-burger-prices-up-14-percent-two-years/814466/",
]

# ── Bottom-up cross-check (rough; only where a seed ingredient clearly matches) ──
GST = 0.10
def inv_lookup(inventory):
    return { r["name"].lower(): float(r["cost"] or 0) for r in inventory if r.get("name") }

# explicit portion assumptions (documented in the caveats sheet)
BUN_BRIOCHE = 0.80          # inventory: Brioche bun $0.80/unit
CUP_LID = 0.30              # inventory: Coffee cup 6oz $6/pack ≈ $0.30 + lid
COFFEE_BEANS = 0.18; MILK = 0.10; TEABAG = 0.15
PATTY_G = 0.120            # 120 g protein portion for burgers
ROLL_PROTEIN_G = 0.090    # 90 g protein for rolls/bagels/toasties
GARNISH_SAUCE = 0.60; CHEESE = 0.30; PACKAGING = 0.30

PROTEIN_KEYWORDS = [   # (keyword in displayName, inventory item name)
    ("moo", "beef brisket"), ("brisket", "beef brisket"), ("beef", "beef brisket"),
    ("slow pork", "pulled pork"), ("pulled pork", "pulled pork"), ("pork", "pork belly (cured)"),
    ("cluck", "chicken thigh (trimmed)"), ("panko", "chicken thigh (trimmed)"), ("tango", "chicken thigh (trimmed)"),
    ("chic", "chicken thigh (trimmed)"), ("chicken", "chicken thigh (trimmed)"),
    ("fish", "white fish fillet"), ("steph", "white fish fillet"),
    ("lamb", "frozen lamb roll"), ("prawn", "prawn 26/30"), ("squid", "squid ring"),
]
TEA_WORDS = ["tea", "chai", "peppermint", "chamomile", "earl grey", "green tea", "lemongrass", "english breakfast"]

def bottom_up(name, category, inv):
    n = (name or "").lower()
    if category == "Coffee & Tea":
        if any(w in n for w in TEA_WORDS):
            return TEABAG + CUP_LID, "tea bag + cup/lid"
        if any(w in n for w in ["choc", "mocha", "chai latte"]):
            return 0.45 + MILK + CUP_LID, "choc/mocha powder + milk + cup"
        return COFFEE_BEANS + MILK + CUP_LID, "beans + milk + cup/lid"
    if category in ("🍔 Burgers", "Rolls", "Bagels", "Toasties"):
        portion = PATTY_G if category == "🍔 Burgers" else ROLL_PROTEIN_G
        for kw, item in PROTEIN_KEYWORDS:
            if kw in n and item in inv:
                base = BUN_BRIOCHE if category in ("🍔 Burgers", "Rolls") else 0.50
                c = inv[item] * portion + base + CHEESE + GARNISH_SAUCE + PACKAGING
                return c, f"{item} {int(portion*1000)}g + base/cheese/sauce/pkg"
        # Burgers: names rarely state the protein → indicative GENERIC beef patty so
        # the category still gets a sanity check (flagged as an assumption).
        if category == "🍔 Burgers" and "beef brisket" in inv:
            c = inv["beef brisket"] * PATTY_G + BUN_BRIOCHE + CHEESE + GARNISH_SAUCE + PACKAGING
            return c, "GENERIC beef patty 120g (protein not named) — indicative only"
        return None, ""
    return None, ""

# ── compute ──
def main():
    menu, menu_path = load_menu()
    inventory = load_inventory()
    inv = inv_lookup(inventory)
    items = menu.get("menuItems", [])

    rows = []
    cat_counts = {}
    for m in items:
        cat = m.get("category", "")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        cls, lo, mid, hi, basis = BENCHMARKS.get(cat, FALLBACK)
        gst_app = m.get("gstApplicable", True) is not False
        sell_ex = float(m.get("sellPrice") or 0)
        sell_inc = round(sell_ex * (1 + GST), 4) if gst_app else sell_ex
        est_cost = round(sell_ex * mid / 100.0, 4)
        est_pct = mid                      # est$ = sell × mid% ⇒ est% ≡ mid
        gm = round(sell_ex - est_cost, 4)
        gm_pct = round(100 - est_pct, 1)

        vrange = ""
        if m.get("hasVariants") and m.get("variants"):
            vps = [float(v.get("sellPrice") or 0) for v in m["variants"]]
            vps = [p for p in vps if p > 0]
            if vps:
                vrange = f"{min(vps):.2f}-{max(vps):.2f}"

        bu, bu_note = bottom_up(m.get("displayName"), cat, inv)
        if bu is not None:
            bu = round(bu, 2)
        bu_pct = round(bu / sell_ex * 100, 1) if (bu is not None and sell_ex > 0) else None
        divergence = round(bu_pct - est_pct, 1) if bu_pct is not None else None

        flags = []
        if sell_ex == 0:
            flags.append("ZERO price (Open Item/placeholder)")
        elif sell_ex < 1.00:
            flags.append("Very low price — cost% volatile")
        if mid > 55:
            flags.append("RESALE high passthrough (>55% benchmark) — verify")
        if bu_pct is not None and bu_pct > 55:
            flags.append("Bottom-up >55% — premium input / possibly underpriced")
        if bu_pct is not None and bu_pct < 10:
            flags.append("Bottom-up <10% — check portion/price")

        rows.append(dict(
            id=m.get("id"), name=m.get("displayName", ""), category=cat, cls=cls,
            sell_ex=sell_ex, sell_inc=sell_inc, lo=lo, mid=mid, hi=hi,
            est_cost=est_cost, est_pct=est_pct, gm=gm, gm_pct=gm_pct,
            vrange=vrange, bu=bu, bu_pct=bu_pct, divergence=divergence,
            bu_note=bu_note, flag="; ".join(flags),
        ))

    rows.sort(key=lambda r: (r["category"], -r["sell_ex"]))
    write_xlsx(rows, cat_counts, menu_path, len(inventory))

# ── xlsx ──
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
FLAG_FILL = PatternFill("solid", fgColor="FDE68A")
RESALE_FILL = PatternFill("solid", fgColor="FEE2E2")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def write_xlsx(rows, cat_counts, menu_path, n_inv):
    wb = Workbook()

    # Sheet 1 — Per item
    ws = wb.active; ws.title = "Per item"
    headers = ["Item ID", "Display name", "Category", "Made / bought", "Sell ex-GST $",
               "Sell inc-GST $", "Bench low %", "Bench mid %", "Bench high %",
               "Est food cost $", "Est food cost %", "Gross margin $", "Margin %",
               "Variant price range ex", "Bottom-up $", "Bottom-up %", "Divergence (pts)",
               "Bottom-up basis", "FLAG"]
    ws.append(headers); style_header(ws, len(headers))
    money = '"$"#,##0.00'; pct = '0.0"%"'
    for r in rows:
        ws.append([
            r["id"], r["name"], r["category"], r["cls"], r["sell_ex"], r["sell_inc"],
            r["lo"], r["mid"], r["hi"], r["est_cost"], r["est_pct"], r["gm"], r["gm_pct"],
            r["vrange"], (r["bu"] if r["bu"] is not None else ""),
            (r["bu_pct"] if r["bu_pct"] is not None else ""),
            (r["divergence"] if r["divergence"] is not None else ""),
            r["bu_note"], r["flag"],
        ])
        rr = ws.max_row
        for col in (5, 6, 10, 12, 15): ws.cell(rr, col).number_format = money
        for col in (7, 8, 9, 11, 13, 16, 17): ws.cell(rr, col).number_format = pct
        if r["flag"]:
            ws.cell(rr, len(headers)).fill = FLAG_FILL
        if r["cls"] == RESALE:
            ws.cell(rr, 4).fill = RESALE_FILL
    widths = [9, 30, 22, 16, 12, 12, 9, 9, 9, 13, 13, 13, 9, 16, 11, 11, 12, 34, 40]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Sheet 2 — Category benchmarks
    ws2 = wb.create_sheet("Category benchmarks")
    h2 = ["Category", "Classification", "Low %", "Mid %", "High %", "# items", "Basis / source"]
    ws2.append(h2); style_header(ws2, len(h2))
    seen = set()
    ordered = sorted(cat_counts.keys(), key=lambda c: (BENCHMARKS.get(c, FALLBACK)[0], c))
    for c in ordered:
        cls, lo, mid, hi, basis = BENCHMARKS.get(c, FALLBACK)
        ws2.append([c, cls, lo, mid, hi, cat_counts[c], basis]); seen.add(c)
        rr = ws2.max_row
        for col in (3, 4, 5): ws2.cell(rr, col).number_format = pct
        ws2.cell(rr, 7).alignment = WRAP
        if cls == RESALE:
            ws2.cell(rr, 2).fill = RESALE_FILL
    for i, w in enumerate([24, 18, 8, 8, 8, 8, 95], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3 — Methodology & caveats
    ws3 = wb.create_sheet("Methodology & caveats")
    ws3.column_dimensions["A"].width = 120
    def line(t, bold=False, size=11):
        ws3.append([t]); c = ws3.cell(ws3.max_row, 1)
        c.font = Font(bold=bold, size=size); c.alignment = WRAP
    line("Mad Benji — Menu food-cost ESTIMATE (planning figures only)", bold=True, size=14)
    line("")
    line("⚠ THESE ARE UNVERIFIED ESTIMATES. Every food-cost figure is derived from category-level industry", bold=True)
    line("benchmarks applied to the sell price — NOT from actual recipes. They must be validated by the chef")
    line("against real recipes/portions/supplier prices BEFORE being used for menu pricing decisions or fed into")
    line("labour, margin, or sales forecasting. Nothing here was written to Firestore; no live cost/recipe field")
    line("was populated.")
    line("")
    line("How each estimate was derived", bold=True, size=12)
    line("1. Each menu CATEGORY was first classified as 'Made in-house' or 'Bought-in / resale', because their")
    line("   cost structures differ sharply (a resold can of soft drink passes through ~40-50% cost; an espresso")
    line("   made in-house is ~20%).")
    line("2. A researched food-cost % range (low / mid / high) was assigned per category from AU 2025-2026 sources.")
    line("3. Per item:  Estimated food cost $ = Sell price (EX-GST) × category MID %.")
    line("   Therefore 'Est food cost %' equals the category mid % by construction; Margin % = 100 − that %.")
    line("4. Inc-GST is shown for reference only (ex × 1.1 where GST applies). All cost math uses EX-GST.")
    line("5. Variant items are costed at the item's headline (default-variant) price; the variant price range is")
    line("   shown separately. Apply the same category % to each variant price for per-size figures.")
    line("")
    line(f"Overall anchor: {OVERALL_NOTE}")
    line("")
    line("Bottom-up cross-check (rough sanity only)", bold=True, size=12)
    line(f"Where a seed inventoryItem ({n_inv} items read from live Firestore, read-only) clearly matches a main")
    line("ingredient, a rough bottom-up cost was computed and compared to the benchmark estimate ('Divergence').")
    line("Portion/consumable assumptions (INDICATIVE — confirm with chef):")
    line("  • Coffee: beans $0.18 + milk $0.10 + takeaway cup/lid $0.30 ≈ $0.58.  Tea: tea bag $0.15 + cup $0.30.")
    line("  • Burger: 120 g matched protein + brioche bun $0.80 + cheese $0.30 + garnish/sauce $0.60 + packaging $0.30.")
    line("  • Roll/bagel/toastie: 90 g matched protein + base $0.50 + cheese/sauce/packaging.")
    line("  • Protein keyword→ingredient matches are heuristic (e.g. 'Mad Moo'→beef brisket, 'Mad Cluck'→chicken thigh).")
    line("  • Bottom-up is left blank where no ingredient clearly matches — most non-burger food has no seed recipe.")
    line("")
    line("Outlier flags", bold=True, size=12)
    line("  • RESALE high passthrough: category benchmark mid > 55% (Retails) — expected for pure resale, but verify item mix.")
    line("  • Bottom-up >55% / <10%: the rough recipe check implies an implausible cost — likely premium input, wrong")
    line("    portion assumption, mispriced, or miscategorised item. Investigate before trusting the price.")
    line("  • Very low / zero price: sub-$1 add-ons (Dips & Sauce) and the $0 'Open Item' placeholder — cost% is")
    line("    meaningless at these prices.")
    line("")
    line("⚠ Explicit assumptions made (each could shift an item by 5-20 pts)", bold=True, size=12)
    for a in [
        "Category benchmarks are GENERIC AU cafe/QSR figures, not Mad Benji's actual supplier costs.",
        "The MID point of each range was used as the point estimate; true cost may sit anywhere in low–high.",
        "'Sparkling Pops' assumed house-made; if packaged/bought-in, its cost% should be resale (35-50%), not 35%.",
        "'Retails' treated as pure resale at 55-75%; the category spans $5-$81 and almost certainly mixes item types.",
        "Bought-in (Pastries/Croissant/Retails/Soft Drinks/Beer) assume standard wholesale markups, not actual invoices.",
        "Bottom-up portions (protein grams, garnish, packaging) are assumed industry-typical, not measured recipes.",
        "GST treated as 10% where gstApplicable; all sell prices taken as EX-GST per the import convention.",
        "Beef-heavy burgers may exceed the 40% mid given 2023-2025 beef cost rises (+32%); treat 40% as a floor.",
        "Labour to produce items (barista/kitchen time) is NOT included — this is FOOD cost only, not prime cost.",
    ]:
        line(f"  • {a}")
    line("")
    line("Sources", bold=True, size=12)
    for s in SOURCES: line(f"  • {s}")
    line("")
    line(f"Source data: {os.path.basename(menu_path)} (281 menuItems) + _inventory_costs.json ({n_inv} inventoryItems, read-only).")
    line("Generated read-only. No Firestore writes. Estimates pending chef sign-off.")

    out = os.path.join(HERE, "MadBenji_Menu_Cost_Estimate.xlsx")
    wb.save(out)
    # console summary
    flagged = sum(1 for r in rows if r["flag"])
    bu_n = sum(1 for r in rows if r["bu"] is not None)
    print(f"Wrote {out}")
    print(f"  rows: {len(rows)} menu items | flagged: {flagged} | bottom-up cross-checks: {bu_n}")
    print(f"  categories: {len(cat_counts)}")

if __name__ == "__main__":
    main()
