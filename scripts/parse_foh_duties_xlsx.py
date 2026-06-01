#!/usr/bin/env python3
"""Parse 'FOH Work Duties and During Service.xlsx' (Sheet1 matrix) into per-role
training modules with Work Duties + During Service sections. Emits JSON with a
top-level `venues` list so the same modules import into multiple venues."""
import sys, json, re
from openpyxl import load_workbook

PATH = sys.argv[1]
VENUES = sys.argv[2].split(",") if len(sys.argv) > 2 else ["hey-sister", "mad-hotpot"]

# role start (number) column index (1-based); text col = +1
ROLES = [
    (1,  "FOH — Bar", "🍹", "Bar (counter & help floor) duties and during-service responsibilities."),
    (11, "FOH — Floor", "🍽️", "Floor service duties and during-service responsibilities."),
    (21, "FOH — Barista", "☕", "Barista station duties and during-service responsibilities."),
    (31, "FOH — Night Bar", "🌙", "Night-time bar (cold drinks & counter service) duties."),
    (41, "FOH — Night Floor", "🌙", "Night-time floor (floor & prep) duties."),
]

def clean(v):
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def main():
    ws = load_workbook(PATH, data_only=True)["Sheet1"]
    modules = []
    for numcol, title, icon, desc in ROLES:
        sections = []
        cur = None
        for r in range(2, ws.max_row + 1):
            left = clean(ws.cell(row=r, column=numcol).value)
            right = clean(ws.cell(row=r, column=numcol + 1).value)
            header = (left or right).upper()
            if header in ("WORK DUTIES", "DURING SERVICE"):
                cur = {"heading": "Work Duties" if "WORK" in header else "During Service", "items": []}
                sections.append(cur)
                continue
            if right and cur is not None:
                cur["items"].append(right)
        sections = [s for s in sections if s["items"]]
        if not sections:
            continue
        modules.append({
            "key": "foh-" + re.sub(r"[^a-z0-9]+", "-", title.lower().replace("foh — ", "")).strip("-") + "-duties",
            "title": title, "cat": "FOH", "duration": "30 min", "icon": icon,
            "color": "#dcfce7", "mandatory": True, "desc": desc, "steps": sections,
        })

    print(json.dumps({"sourceTitle": "FOH Work Duties and During Service.xlsx",
                      "venues": VENUES, "modules": modules}, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
